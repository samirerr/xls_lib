#!/usr/bin/python

from __future__ import absolute_import, division, print_function

__metaclass__ = type

# import modules
import os
import openpyxl
from ansible.module_utils.basic import AnsibleModule
from openpyxl import Workbook


DOCUMENTATION = r"""
----
module: xls_write
Version: 1.0
short_description: Write facts to excel spreadsheet
description:
- This module writes facts into a excel spreadsheet. The name of the spreadsheet 
    can be specified.
- The facts needs to be a list of dictonaries. It uses the key in the first item 
    in the list as headers to the excel spreadsheet.
- In the current state, it always returns as state changed because if it detects 
    that the workbook already has a spreadsheet with the same name, it will delete 
    the spreadsheet and create a new one with the new data.
- #TODO: Add support for checkmode and diff mode.
options:
    path:
        description: The path to the file that needs to be modified.
        type: path
        required: True
    workbook:
        description: The name of the excel spreadsheet that needs to be modified.
        type: str
        required: True
    worksheet: 
        description: The name of the work sheet that needs to be modified.
        type: str
        required: True
    data:
        description: The actual facts that need to be written in the file.
        type: list
        required: True
    create:
        description: If specified, the file will be created if it does not exist.
        type: bool
        required: False
        default: False
    create_header:
        description: If specified, it will create a header of the excel excel file.
        type: bool
        required: False
        default: False
    headers:
        description: it's a list which containt the names of headers .
        type: bool
        required: False
        default: False
"""

EXAMPLES = r"""

    # Example data without creating header
    vars:
        create_header: no
        data:
            - header1: value1
              header2: value2
              header3: value3
              
            - header1: another_value1
              header2: another_value2
              header3: another_value3
    # Example data with reating header and insert data into appropriate header by key
    vars:
        create_header: yes
        headers: 
            - header1
            - header2
            - header3
        data:
            - header1: value1
              header2: value2
              header3: value3
    #! Note: it is important and necessary to include the .xlsx extention in the
    #!       workbook name.
    - name: Write facts to spreadsheet
      xls_write:
        path: ./result
        workbook: workbook.xlsx
        worksheet: worksheet
        data: "{{ data_list }}"
        create: yes
        create_header: yes
"""



def write_xls(module,dest, create, workbook, worksheet,create_header, headers, data):
    book_created = False
    changed = False
    if not os.path.exists(dest):
        if not create:
            module.fail_json(msg=f"Destination folder '{dest}' does not exist"+ "! Set create to 'True' to create the"+ "destination folder.")
            return 
        else: 
            try:
                os.makedirs(dest)
                book_created = True
                changed = True
            except Exception as err:
                module.fail_json(msg=f"Error creating {dest} ({err})")
    else:
        if not os.path.isfile(f"{dest}/{workbook}"):
            if not create:
                module.fail_json(msg=f"Workbook '{workbook}' does not exist"+ "! Set Create to 'True' to create"+ "a new workbok.")
                return 
            else: 
                try:
                    book = Workbook()
                    changed = True
                    book_created= True

                except Exception as err:
                    module.fail_json(msg=f"Error creating {workbook} ({err})")
        else:
            try:
                book = openpyxl.load_workbook(f"{dest}/{workbook}", data_only=True)
                work_sheet = book.active
                create_header = False
            except Exception as err:
                module.fail_json(msg=f"Error creating {workbook} ({err})")
            
    

    work_sheet = book.active
    work_sheet.title = worksheet

    if create_header:
        if len(headers) > 0  :
            try:
                work_sheet.append(headers)    
            except Exception as err:
                print(f"{err}: headers must be a list .")
                #module.fail_json(msg=f"{err}: headers must be a list .")
        else:
            print(f"{err}: headers must not be empty.")
            #module.fail_json(msg=f"{err}: headers must not be empty.")
    
        for entry in data:
            data_write = []
            row_no = work_sheet.max_row + 1 
            for key, value in entry.items():
                data_write.append(str(value))
                pos= headers.index(key)+1
                work_sheet.cell(row_no,pos).value = value      
    else:
        for entry in data:
            data_write = []
            for key, value in entry.items():
                data_write.append(str(value))
            work_sheet.append(data_write)
           

    # work_sheet.cell(row=5, column=5).value = 2
    book.save(f"{dest}/{workbook}")


    
    module.exit_json(changed=changed, msg="Done!")


def main():
    # Define Module Parameters
    module = AnsibleModule(
        argument_spec=dict(
            path=dict(type="path", required=True),
            workbook=dict(type="str", required=True),
            worksheet=dict(type="str", required=True),
            data=dict(type="list", required=True),
            headers=dict(type="list", default=False),
            create=dict(type="bool", default=False),
            create_header=dict(type="bool", default=False),
        )
    )
    # Assign Module parameters
    params = module.params
    path = params["path"]
    workbook = params["workbook"]
    worksheet = params["worksheet"]
    data = params["data"]
    headers = params["headers"]
    create = params["create"]
    create_header = params["create_header"]
    # Ensure that the required parameters are passed and error out with an
    # appropriate message if they are not.
    if path is None:
        module.fail_json(msg="path is required")
    if workbook is None:
        module.fail_json(msg="workbook is required")
    if worksheet is None:
        module.fail_json(msg="worksheet is required")
    if data is None:
        module.fail_json(msg="data is required")
    # Call function that handles the write operation
    write_xls(module, path, create, workbook, worksheet,create_header,headers, data)


# Invoke main function.
if __name__ == "__main__":
    main()